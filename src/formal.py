import openpyxl as op
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment
import ops_defines as od
import os

output_name = "clans_infor.xlsx"

def clear_sheet_all(file_name, sheet_name):
    """
    清除指定工作表的所有内容、格式和合并单元格，完全重置工作表。
    
    :param file_name: Excel 文件的名称
    :param sheet_name: 目标工作表的名称
    """
    # 加载工作簿
    wb = op.load_workbook(file_name)
    
    # 获取目标工作表
    ws = wb[sheet_name]
    # 删除所有合并单元格
    ws.merged_cells = []  # 直接清空合并单元格
    # 取消所有合并单元格
    for merged_range in list(ws.merged_cells):
        ws.unmerge_cells(str(merged_range))  # 取消合并
    
    # 清除所有单元格的内容和格式
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None  # 清空内容
            cell.style = 'Normal'  # 清除格式
    # 保存修改后的工作簿
    wb.save(file_name)

def check_file(file_name):##当前目录下是否存在输出workbook
    if os.path.isfile(file_name):
        return True
    else:
        return False
    
def del_sheet_from_wb(file_name,sheet_name):##特定的sheet是否存在
    wb = op.load_workbook(filename = file_name)
    if sheet_name in wb.sheetnames :##存在即删除
        if len(wb.sheetnames) > 1:##删或者清除
            del wb[sheet_name]
        else :
            clear_sheet_all(file_name,sheet_name)
    wb.save(filename = file_name)
    ###不存在无事

def creat_sheet(op_num):
    if od.creat_contribution_sheet() == op_num:
        sheet_name = "Contribution"
        if check_file(output_name):##输出文件是否存在
            ###存在则删除贡献sheet
            del_sheet_from_wb(output_name,sheet_name)
            wb = op.load_workbook(output_name)
            wb.create_sheet(title = sheet_name)
        else :##不存在新建
            wb = op.Workbook()
            wb.create_sheet(title = sheet_name)
            # if len(wb.sheetnames) > 1:##删掉不为空
            #     del wb['Sheet']##删掉Sheet
        ws = wb[sheet_name]
        ws.append(['部落','玩家','今日使用卡组数','袭击战船次数','总贡献'])##表头
        wb.save(filename = output_name)
    else :
        print("Error!")

def calculate_adjusted_width(value):
    """
    根据单元格内容计算宽度，处理中文繁体和宽字符的显示宽度。
    :param value: 单元格内容
    :return: 调整后的宽度（适配 Excel 的列宽标准）
    """
    if value is None:
        return 0
    value = str(value)
    width = 0
    for char in value:
        # 宽字符 (中文、繁体、日文、韩文、emoji等)
        if '\u4e00' <= char <= '\u9fff' or ord(char) > 127:
            width += 2  # 宽字符占用2单位宽度
        else:
            width += 1  # 普通字符占用1单位宽度
    return width

def adjust_column_width(sheet):
    """
    动态调整工作表的列宽，确保中文繁体和宽字符正确显示。
    :param sheet: 当前工作表
    """
    for col in sheet.columns:
        max_width = 0
        col_letter = get_column_letter(col[0].column)  # 获取列号字母
        for cell in col:
            if cell.value is not None:
                # 根据内容计算宽度
                adjusted_width = calculate_adjusted_width(cell.value)
                max_width = max(max_width, adjusted_width)
        # Excel的列宽与字符宽度不同，这里乘1.2是一个经验调整值
        sheet.column_dimensions[col_letter].width = max_width * 1.2

def process_excel(file_name):
    """
    处理Excel文件：
    1. 检查当前目录是否存在指定的Excel文件。
       - 如果不存在，抛出错误提示。
    2. 删除名为 "Sheet" 的工作表（如果 sheetnames 数量大于1）。
    3. 对剩下的每个工作表：
        - 调整列宽。
        - 设置单元格内容居中。
    :param file_name: str, Excel 文件名。
    """
    # 检查当前目录是否存在指定文件
    if not os.path.exists(file_name):
        print(f"文件 '{file_name}' 不存在！请检查文件名和路径。")
        return

    # 加载工作簿
    wb = op.load_workbook(file_name)

    # 删除 "Sheet" 工作表
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 遍历每个工作表
    for sheet in wb.worksheets:
        # 调整列宽
        adjust_column_width(sheet)

        # 设置单元格居中
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 保存更改
    wb.save(file_name)